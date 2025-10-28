import os
import json
import openpyxl
from pathlib import Path
import calendar

# Define directories
INPUT_DIR = "./input_data"
OUTPUT_DIR = "./transformed_data/sharepoint_excel_to_json_data"

# Valid month names for validation
VALID_MONTHS = [month.lower() for month in calendar.month_name if month] + \
               [month.lower() for month in calendar.month_abbr if month]

def is_valid_month_sheet(sheet_name):
    """Check if sheet name contains a valid month name."""
    sheet_lower = sheet_name.lower()
    return any(month in sheet_lower for month in VALID_MONTHS)

def get_last_populated_column(sheet, row_num):
    """Find the last populated column in a given row."""
    max_col = sheet.max_column
    for col in range(max_col, 0, -1):
        cell_value = sheet.cell(row=row_num, column=col).value
        if cell_value is not None and str(cell_value).strip():
            return col
    return None

def extract_team_members(sheet):
    """Extract team member names from column A until 'Sprint No.' or empty cell."""
    team_members = []
    row = 1
    sprint_start_row = None
    
    while True:
        cell_value = sheet.cell(row=row, column=1).value
        
        # Check if we encounter "Sprint No."
        if "sprint no" in str(cell_value).lower():
            sprint_start_row = row
            break
        
        # Add non-empty team member name
        if cell_value and str(cell_value).strip().lower() not in ["name"]:
            team_members.append((str(cell_value).strip(), row))
        
        row += 1
    
    return team_members, sprint_start_row

def extract_sprint_info(sheet, sprint_start_row):
    """Extract sprint information starting from the Sprint No. row."""
    if sprint_start_row is None:
        return {}
    
    sprint_info = {}
    row = sprint_start_row + 1
    
    while True:
        cell_value = sheet.cell(row=row, column=1).value
        
        # Stop if empty cell
        if cell_value is None or str(cell_value).strip() == "":
            break
        
        sprint_number = str(cell_value).strip()
        
        # Find sprint name and URL in the row
        sprint_url = str(sheet.cell(row=row, column=2).value).strip()
        sprint_name = str(sheet.cell(row=row, column=3).value).strip()
        
        # Create sprint key (sprint_1, sprint_2, etc.)
        sprint_key = f"sprint_{sprint_number}" if sprint_number.isdigit() else sprint_number
        
        sprint_info[sprint_key] = {
            "name_of_sprint": sprint_name,
            "url": sprint_url
        }
        
        row += 1
    
    return sprint_info

def extract_row_data(sheet, row_num):
    """Extract all data from a row until an empty column is encountered."""
    data = []
    col = 1
    
    while True:
        header_cell = sheet.cell(row=1, column=col).value
        value_cell = sheet.cell(row=row_num, column=col).value
        
        # Stop if we encounter an empty header
        if header_cell is None or str(header_cell).strip() == "":
            break
        
        header = str(header_cell).strip()
        value = str(value_cell).strip() if value_cell is not None else ""
        
        # Append specific columns only
        if any(header_name_to_filter in header.lower() for header_name_to_filter in ["sprint commitments", 
                                                                                     "mini quizzes", 
                                                                                     "monthly evaluation", 
                                                                                     "final evaluation",
                                                                                     "hackathon",
                                                                                     "total score"]):
            data.append((header, value))
        
        col += 1
    
    return data

def process_excel_file(file_path):
    """Process a single Excel file and return structured data."""
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    result = {}
    
    for sheet_name in workbook.sheetnames:
        # Skip sheets without valid month names
        if not is_valid_month_sheet(sheet_name):
            continue
        
        sheet = workbook[sheet_name]
        sheet_data = {}
        
        # Extract team members and their row numbers, and sprint start row
        team_members, sprint_start_row = extract_team_members(sheet)
        
        for member_name, row_num in team_members:
            # Extract all data for this team member
            member_data = extract_row_data(sheet, row_num)
            sheet_data[member_name] = member_data
        
        # Extract sprint information
        sprint_info = extract_sprint_info(sheet, sprint_start_row)
        if sprint_info:
            sheet_data["sprint_info"] = sprint_info
        
        result[sheet_name] = sheet_data
    
    workbook.close()
    return result

def main():
    """Main function to process all Excel files."""
    # Create output directory if it doesn't exist
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    
    # Process each .xlsx file in input directory
    if not os.path.exists(INPUT_DIR):
        print(f"Error: Input directory '{INPUT_DIR}' does not exist.")
        return
    
    xlsx_files = [f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx')]
    
    if not xlsx_files:
        print(f"No .xlsx files found in '{INPUT_DIR}'.")
        return
    
    for filename in xlsx_files:
        input_path = os.path.join(INPUT_DIR, filename)
        output_filename = filename.replace('.xlsx', '.json')
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        print(f"Processing: {filename}")
        
        try:
            # Process the Excel file
            data = process_excel_file(input_path)
            
            # Write to JSON file
            with open(output_path, 'w', encoding='utf-8') as json_file:
                json.dump(data, json_file, indent=4, ensure_ascii=False)
            
            print(f"  ✓ Generated: {output_filename}")
        
        except Exception as e:
            print(f"  ✗ Error processing {filename}: {str(e)}")
    
    print("\nTransformation complete!")

if __name__ == "__main__":
    main()
